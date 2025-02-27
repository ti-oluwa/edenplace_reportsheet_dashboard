import enum
import math
from pathlib import Path
import itertools
import typing

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def remove_empty_first_rows(worksheet: Worksheet):
    """Remove empty rows from the beginning of the worksheet."""
    row_idx = 1
    # Loop until the first non-empty row is found
    while row_idx <= worksheet.max_row:
        # Check if all cells in the row are empty
        row = worksheet[row_idx]
        if all(cell.value is None for cell in row):
            # If the row is empty, delete it
            worksheet.delete_rows(row_idx)
        else:
            # If a non-empty row is found, stop deleting
            break

    return worksheet


def load_workbook(path: Path):
    """Load an Excel workbook from a file path."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    return workbook


def nonempty_worksheets(workbook: openpyxl.Workbook):
    """Yield non-empty worksheets from a workbook."""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        if worksheet.max_row < 5 or worksheet.max_column < 3:
            # This ensure that empty worksheets are not processed
            # As the processing of empty worksheets will take a lot of time
            # because of the schema extraction process will have to traverse
            # the entire worksheet.
            continue
        yield worksheet


Term = typing.Type[str]
"""Type alias for a school term name."""
AggregateName = typing.Type[str]
"""Type alias for an result aggregate name."""
SubjectName = typing.Type[str]
"""Type alias for a subject name."""


class SchemaInfo(typing.TypedDict):
    """Schema information for a column in a worksheet."""

    column: int
    overall: typing.NotRequired[typing.Optional[int]]


class SubjectSchema(typing.TypedDict):
    """Schema information for a subject in a worksheet."""

    mid_term_score: SchemaInfo
    exam_score: SchemaInfo
    total_score: SchemaInfo


SubjectsSchemas = typing.Dict[SubjectName, SubjectSchema]
"""Type alias for a collection or mapping of subject schemas."""
AggregatesSchemas = typing.Dict[AggregateName, SchemaInfo]
"""Type alias for a collection or mapping of aggregate schemas."""


class BroadSheetSchema(typing.TypedDict):
    """Schema information for a broadsheet worksheet."""

    term: Term
    subjects: SubjectsSchemas
    aggregates: AggregatesSchemas
    teachers_comment: SchemaInfo
    coordinators_comment: SchemaInfo


EXTERNAL_TO_INTERNAL_MAPPING = {
    "mid": "mid_term_score",
    "exam": "exam_score",
    "total": "total_score",
    "mid %": "mid term %",
    "mid total": "mid term total",
    "sim %": "sum total %",
    "sum %": "sum total %",
    "1st term": "1st term total",
    "2nd term": "2nd term total",
    "3rd term": "3rd term total",
    "cumtotal": "cumulative (session) total",
    "av. total": "average total",
    "av. %": "average %",
}
"""Mapping of all possible (external) column names - from the broadsheet, 
to internal names - to be used in the code, and report generation."""


def _to_internal(val: str) -> str:
    """Convert an external column name to an internal column name."""
    val = val.strip().lower()
    return EXTERNAL_TO_INTERNAL_MAPPING.get(val, val)


def get_broadsheet_schema(worksheet: Worksheet):
    """
    Extract the schema information from a broadsheet worksheet.

    :param worksheet: The worksheet/broadsheet whose schema is to be extracted.
    """
    # Use typed dict for detailed typing and dictionary data access
    schema = BroadSheetSchema(
        term=worksheet.title.strip().title(),
        subjects={},
        aggregates={},
        teachers_comment={},
        coordinators_comment={},
    )
    last_column_index = None

    # Split the columns into batches of 3, because most columns (in row 2) have 3 sub columns (in row 3).
    # We limit the scope we need to iterate over to (row2, column3) to (row3, col*), as that is
    # the cell range in which the schema data we need to extract lies.
    # Simply put, just the 2nd and 3rd row are what we need to extract the column schema schema.
    for cols in itertools.batched(
        worksheet.iter_cols(min_col=3, max_col=300, min_row=2, max_row=4), n=3
    ):
        previous_title = None
        for col in cols:
            title = col[0].value
            sub_title = col[1].value
            sub_title_column_index = col[1].column
            overall = col[2].value

            if title:
                title = _to_internal(title)
                if title not in schema["subjects"]:
                    schema["subjects"][title] = {}
            else:
                if not previous_title:
                    if not sub_title:
                        continue

                    sub_title = _to_internal(sub_title)
                    if "comment" in sub_title:
                        continue

                    if sub_title not in schema["aggregates"]:
                        schema["aggregates"][sub_title] = {}

                    schema["aggregates"][sub_title]["column"] = sub_title_column_index
                    schema["aggregates"][sub_title]["overall"] = overall
                    last_column_index = sub_title_column_index
                    continue
                else:
                    title = previous_title

            if not sub_title:
                continue

            sub_title = _to_internal(sub_title)
            schema["subjects"][title][sub_title] = {"column": sub_title_column_index}
            schema["subjects"][title][sub_title]["overall"] = overall
            last_column_index = sub_title_column_index
            previous_title = title

    if last_column_index is not None:
        # Comments are anticipated to always be in the last column
        schema["teachers_comment"]["column"] = last_column_index + 1
        schema["coordinators_comment"]["column"] = last_column_index + 2
    return schema


class StudentInfo(typing.TypedDict):
    """Information about a student in a worksheet."""

    name: str
    row: int


def students(worksheet: Worksheet):
    """Yield student information from a worksheet/broadsheet."""
    for row in worksheet.iter_rows(min_row=5, min_col=2, max_col=2):
        name = row[0].value
        if not name:
            continue
        yield StudentInfo(
            name=name.strip().title(),
            row=row[0].row,
        )


Value = typing.Optional[typing.Union[int, float]]
"""Type alias for a score value which can be an integer or a float."""


class Grade(enum.StrEnum):
    """Enumeration of possible grades for a subjects overall score."""

    A = "A"
    B = "B"
    C = "C"
    D = "D"
    E = "E"
    F = "F"


def get_grade(score: Value) -> typing.Optional[Grade]:
    """Deduce the grade from an overall score."""
    if score is None:
        return None

    score = round(score, ndigits=1)
    if score >= 85:
        return Grade.A
    elif 70 <= score <= 84:
        return Grade.B
    elif 55 <= score <= 69:
        return Grade.C
    elif 50 <= score <= 54:
        return Grade.D
    elif 45 <= score <= 49:
        return Grade.E
    else:
        return Grade.F


class SubjectScore(typing.TypedDict):
    """Scores for each term section for a subject."""

    mid_term_score: Value
    exam_score: Value
    total_score: Value
    grade: Grade


SubjectsScores = typing.Dict[str, SubjectScore]
"""Type alias for a collection or mapping of subject scores."""
AggregatesValues = typing.Dict[str, Value]
"""Type alias for a collection or mapping of aggregate values."""
StudentName = typing.Type[str]
"""Type alias for a student's name."""


class StudentResult(typing.TypedDict):
    """Result information for a student in a worksheet."""

    term: Term
    student: StudentName
    subjects: SubjectsScores
    aggregates: AggregatesValues
    teachers_comment: typing.Optional[str]
    coordinators_comment: typing.Optional[str]


def get_subjects_scores_for_student(
    worksheet: Worksheet,
    student_row_index: int,
    subjects_schemas: typing.Dict[str, SubjectSchema],
):
    """
    Extract and return the subjects scores for a student in a worksheet.

    :param worksheet: The worksheet/broadsheet containing the student's scores.
    :param student_row_index: The index of the row containing the student's info
        in the worksheet/broadsheet.
    :param subjects_schemas: The schema information for the subjects in the
        worksheet/broadsheet to be used to extract the scores.
    :return: The subjects scores for the student.
    """
    subjects_scores: SubjectsScores = {}
    for subject, subject_schema in subjects_schemas.items():
        mid_term_score_column_index = subject_schema["mid_term_score"]["column"]
        exam_score_column_index = subject_schema["exam_score"]["column"]
        total_score_column_index = subject_schema["total_score"]["column"]
        mid_term_score = worksheet.cell(
            student_row_index, mid_term_score_column_index
        ).value
        exam_score = worksheet.cell(student_row_index, exam_score_column_index).value
        total_score = worksheet.cell(student_row_index, total_score_column_index).value
        subject_score = SubjectScore(
            mid_term_score=mid_term_score,
            exam_score=exam_score,
            total_score=total_score,
            grade=get_grade(total_score),
        )
        subjects_scores[subject] = subject_score
    return subjects_scores


def get_aggregates_values(
    worksheet: Worksheet,
    student_row_index: int,
    aggregates_schemas: typing.Dict[str, SchemaInfo],
):
    """
    Extract and return the aggregate values for a student in a worksheet.

    :param worksheet: The worksheet/broadsheet containing the student's scores.
    :param student_row_index: The index of the row containing the student's info
        in the worksheet/broadsheet.
    :param aggregates_schemas: The schema information for the aggregates in the
        worksheet/broadsheet to be used to extract the values.
    :return: The aggregate values for the student.
    """
    aggregates_values: AggregatesValues = {}
    for aggregate, aggregate_schema in aggregates_schemas.items():
        aggregate_column_index = aggregate_schema["column"]
        aggregate_value = worksheet.cell(
            student_row_index, aggregate_column_index
        ).value
        aggregates_values[aggregate] = (
            round(aggregate_value, ndigits=1) if aggregate_value else None
        )
    return aggregates_values


def get_comment_value(
    worksheet: Worksheet, student_row_index: int, comment_column_index: int
):
    """
    Extract and return the comment value for a student in a worksheet.

    :param worksheet: The worksheet/broadsheet containing the student's scores.
    :param student_row_index: The index of the row containing the student's info
        in the worksheet/broadsheet.
    :param comment_column_index: The index of the column containing the comment
        in the worksheet/broadsheet.
    :return: The comment value for the student.
    """
    comment = worksheet.cell(student_row_index, comment_column_index).value
    if not comment:
        return None
    return comment.strip()


def student_results(
    worksheet: Worksheet, broadsheet_schema: typing.Optional[BroadSheetSchema] = None
):
    """
    Extract and yield the results for each student in a worksheet.

    :param worksheet: The worksheet/broadsheet containing the students' scores.
    :param broadsheet_schema: The schema information for the worksheet/broadsheet.
        to be used to extract the results.
    :return: The results for each student in the worksheet/broadsheet.
    """
    broadsheet_schema = broadsheet_schema or get_broadsheet_schema(worksheet)
    for student in students(worksheet):
        student_row_index = student["row"]
        student_name = student["name"]
        subjects_schemas = broadsheet_schema["subjects"]
        aggregates_schemas = broadsheet_schema["aggregates"]
        teachers_comment_schema = broadsheet_schema["teachers_comment"]
        coordinators_comment_schema = broadsheet_schema["coordinators_comment"]

        subjects_scores = get_subjects_scores_for_student(
            worksheet=worksheet,
            student_row_index=student_row_index,
            subjects_schemas=subjects_schemas,
        )
        aggregates_values = get_aggregates_values(
            worksheet=worksheet,
            student_row_index=student_row_index,
            aggregates_schemas=aggregates_schemas,
        )
        teachers_comment = get_comment_value(
            worksheet=worksheet,
            student_row_index=student_row_index,
            comment_column_index=teachers_comment_schema["column"],
        )
        coordinators_comment = get_comment_value(
            worksheet=worksheet,
            student_row_index=student_row_index,
            comment_column_index=coordinators_comment_schema["column"],
        )
        result = StudentResult(
            term=broadsheet_schema["term"],
            student=student_name,
            subjects=subjects_scores,
            aggregates=aggregates_values,
            teachers_comment=teachers_comment,
            coordinators_comment=coordinators_comment,
        )
        yield result


class BroadSheetData(typing.TypedDict):
    """Data extracted from a broadsheet/worksheet."""

    students_results: typing.List[StudentResult]
    broadsheet_schema: BroadSheetSchema


BroadSheetsData = typing.Dict[Term, BroadSheetData]
"""Type alias for a collection or mapping of broadsheet data."""


def extract_broadsheets_data(file: typing.Union[str, Path]):
    """
    Extract and return the data from a file containing broadsheets (excel workbook).

    :param file: The path to the file containing the broadsheets.
    :return: The data extracted from the broadsheets.
    """
    workbook = load_workbook(Path(file).resolve())
    broadsheets_data: BroadSheetsData = {}
    for worksheet in nonempty_worksheets(workbook):
        worksheet = remove_empty_first_rows(worksheet)
        broadsheet_schema = get_broadsheet_schema(worksheet)

        results: typing.List[StudentResult] = []
        for student_result in student_results(
            worksheet, broadsheet_schema=broadsheet_schema
        ):
            results.append(student_result)

        term = broadsheet_schema["term"]
        broadsheets_data[term] = BroadSheetData(
            students_results=results,
            broadsheet_schema=broadsheet_schema,
        )
    return broadsheets_data
